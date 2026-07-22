"""json → xlsx → json must not lose or corrupt data.

The xlsx carries no `sources` and no `$comment` column, so those two are the
only legitimate losses; everything else must survive byte-identical.
"""
import json

import openpyxl
import pytest

from profile_json_to_xlsx import build
from xlsx_to_profile_json import extract, split_name, strip_ts_label

LOSSY_KEYS = {"sources", "$comment"}


def roundtrip(profile, tmp_path, name="rt.xlsx"):
    path = tmp_path / name
    build(profile).save(path)
    return extract(str(path))


def test_all_fields_survive(profile, tmp_path):
    back = roundtrip(profile, tmp_path)
    for key in set(profile) - LOSSY_KEYS:
        assert back.get(key) == profile[key], f"{key} 在來回轉換中改變了"


def test_single_row_sections_survive(profile, tmp_path):
    """A one-entry section produces a 1x1 merged cell (e.g. A9:A9); the reader
    finds sections by scanning merged ranges, so this is the case that would
    break first if openpyxl stopped recording single-cell merges."""
    profile["education"] = [{"school": "某大", "major": "資工",
                             "degree_level": "學士", "degree": "工學士"}]
    profile["positions"] = ["某公司 執行長"]
    profile["career"] = [{"date": "2020-", "org": "某公司", "role": "執行長"}]

    back = roundtrip(profile, tmp_path)
    assert back["education"] == profile["education"]
    assert back["positions"] == profile["positions"]
    assert back["career"] == profile["career"]


def test_empty_sections_come_back_empty(profile, tmp_path):
    profile["education"] = []
    profile["positions"] = []
    profile["career"] = []
    back = roundtrip(profile, tmp_path)
    assert back["education"] == []
    assert back["positions"] == []
    assert back["career"] == []


def test_timestamp_label_does_not_accrete(profile, tmp_path):
    """The writer prefixes A1 with 資料彙整：; the reader must strip it, or each
    pass adds another copy."""
    once = roundtrip(profile, tmp_path, "a.xlsx")
    twice = roundtrip(once, tmp_path, "b.xlsx")
    thrice = roundtrip(twice, tmp_path, "c.xlsx")
    assert thrice["timestamp"] == profile["timestamp"]


def test_name_en_splits_back_out(profile, tmp_path):
    profile["name"] = "王小明"
    profile["name_en"] = "Wang Xiaoming"
    back = roundtrip(profile, tmp_path)
    assert back["name"] == "王小明"
    assert back["name_en"] == "Wang Xiaoming"


def test_chinese_parenthetical_stays_in_name(profile, tmp_path):
    """Only a Latin parenthetical is a name_en; a Chinese aside is part of the
    name and must not be torn off."""
    profile["name"] = "王小明（範例）"
    profile.pop("name_en", None)
    back = roundtrip(profile, tmp_path)
    assert back["name"] == "王小明（範例）"
    assert "name_en" not in back


def test_position_cell_with_a_number_reads_back_as_string(profile, tmp_path):
    """Regression: positions was the one field not passed through cell_value(),
    so a year typed into the position cell came back as an int and failed
    schema validation."""
    path = tmp_path / "numeric.xlsx"
    build(profile).save(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "現任職位":
            ws.cell(row=row, column=2).value = 2026
            break
    wb.save(path)

    back = extract(str(path))
    assert back["positions"][0] == "2026"
    assert all(isinstance(p, str) for p in back["positions"])


@pytest.mark.parametrize("written", ["－", "—", "–", "", None])
def test_empty_spellings_normalise_to_the_sanctioned_one(profile, tmp_path, written):
    """Seed files spell "empty" several ways; reading must collapse them all to
    "-" or the file fails validation on the way back out."""
    path = tmp_path / "dashes.xlsx"
    build(profile).save(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws["B4"] = written           # 出生年月
    wb.save(path)

    assert extract(str(path))["birth"] == "-"


def test_note_survives(profile, tmp_path):
    profile["note"] = "測試備註：部分內容為推估。"
    assert roundtrip(profile, tmp_path)["note"] == profile["note"]


def test_row_count_scales_with_data(profile, tmp_path):
    """Row counts are computed from the JSON, not hardcoded — a longer career
    must produce a taller sheet."""
    short = dict(profile, career=profile["career"][:2])
    long = dict(profile, career=profile["career"][:5])

    short_path, long_path = tmp_path / "s.xlsx", tmp_path / "l.xlsx"
    build(short).save(short_path)
    build(long).save(long_path)

    rows_s = openpyxl.load_workbook(short_path).active.max_row
    rows_l = openpyxl.load_workbook(long_path).active.max_row
    assert rows_l - rows_s == 3


# --- 純函式 -----------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("王小明（Wang Xiaoming）", ("王小明", "Wang Xiaoming")),
    ("王小明", ("王小明", None)),
    ("王小明（範例）", ("王小明（範例）", None)),
])
def test_split_name(raw, expected):
    assert split_name(raw) == expected


def test_strip_ts_label_removes_repeats():
    assert strip_ts_label("資料彙整：資料彙整：2026/01/01") == "2026/01/01"
