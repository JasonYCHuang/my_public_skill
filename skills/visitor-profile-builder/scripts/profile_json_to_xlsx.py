#!/usr/bin/env python3
"""
Render a profile.json (see assets/profile.schema.json) into a
"個人信息登記表"-style xlsx, matching the layout/merged-cell structure of
reference/*.xlsx. Row counts for 教育經歷／現任職位／主要履歷 are computed
from however many entries are in the JSON (not hardcoded to 2/3/7 rows).

The record is a closed set of 10 fields shared with the HTML card — see
references/field-contract.md. Fields with no data hold a literal "-".

Usage:
    python3 profile_json_to_xlsx.py <profile.json> -o <output.xlsx>

Embedding photos requires Pillow (pip install Pillow) in addition to
openpyxl. If Pillow isn't installed, photos are skipped with a warning
(the xlsx still builds fine without them).
"""
import argparse
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from validate_profile import validate_or_exit

LABEL_FILL = PatternFill("solid", fgColor="DCE6F1")
SUBHEAD_FILL = PatternFill("solid", fgColor="F2F2F2")
TITLE_FONT = Font(name="Arial", size=20, bold=True)
LABEL_FONT = Font(name="Arial", size=14, bold=True)
VALUE_FONT = Font(name="Arial", size=14, bold=False)
SUBHEAD_FONT = Font(name="Arial", size=14, bold=True)
DATA_FONT = Font(name="Arial", size=12, bold=False)
NOTE_FONT = Font(name="Arial", size=11, italic=True, color="808080")
TS_FONT = Font(name="Arial", size=10)

thin = Side(style="thin", color="999999")
BORDER_ALL_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

PLACEHOLDER = "-"  # the one sanctioned "no data" marker; see field-contract.md


def val(v):
    return v if v not in (None, "") else PLACEHOLDER


def build(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col, w in {"A": 12.5, "B": 23.0, "C": 18.625, "D": 13.5, "E": 15.125, "F": 27.625}.items():
        ws.column_dimensions[col].width = w
    for c in range(7, 14):
        ws.column_dimensions[get_column_letter(c)].width = 10.0

    def set_cell(coord, value, font=DATA_FONT, fill=None, align=LEFT, border=BORDER_ALL_THIN):
        cell = ws[coord]
        cell.value = value
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = align
        cell.border = border
        return cell

    def merge(rng):
        ws.merge_cells(rng)

    merge("A1:F1")
    set_cell("A1", f'資料彙整：{data.get("timestamp", "")}', font=TS_FONT, align=RIGHT, border=Border())

    merge("A2:F2")
    set_cell("A2", "個人信息登記表", font=TITLE_FONT, align=CENTER, border=Border())
    merge("G2:M3")
    set_cell("G2", "", align=CENTER, border=BORDER_ALL_THIN)

    set_cell("A3", "姓　　名", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    name = data.get("name", "")
    if data.get("name_en"):
        name = f'{name}（{data["name_en"]}）'
    set_cell("B3", name, font=VALUE_FONT, align=LEFT)
    set_cell("C3", "性　　別", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    merge("D3:E3")
    set_cell("D3", val(data.get("gender")), font=VALUE_FONT, align=LEFT)

    edu_rows = data.get("education") or []
    n_edu = max(len(edu_rows), 1)
    edu_header_row = 6
    edu_last_row = edu_header_row + n_edu

    merge(f"F3:F{edu_last_row}")
    set_cell("F3", "（照片詳右方欄位，\n資料彙整自公開網路資訊）", font=NOTE_FONT, align=CENTER)

    set_cell("A4", "出生年月", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    set_cell("B4", val(data.get("birth")), font=VALUE_FONT, align=LEFT)
    set_cell("C4", "生　　肖", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    merge("D4:E4")
    set_cell("D4", val(data.get("zodiac")), font=VALUE_FONT, align=LEFT)

    set_cell("A5", "聯繫方式", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    set_cell("B5", val(data.get("contact")), font=VALUE_FONT, align=LEFT)
    set_cell("C5", "籍　　貫", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    merge("D5:E5")
    set_cell("D5", val(data.get("hometown")), font=VALUE_FONT, align=LEFT)

    merge(f"A{edu_header_row}:A{edu_last_row}")
    set_cell(f"A{edu_header_row}", "教育經歷", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    for col, label in zip("BCDE", ["畢業院校", "專業", "學歷", "學位"]):
        set_cell(f"{col}{edu_header_row}", label, font=SUBHEAD_FONT, fill=SUBHEAD_FILL, align=CENTER)
    r = edu_header_row + 1
    rows_to_write = edu_rows if edu_rows else [{}]
    for e in rows_to_write:
        set_cell(f"B{r}", val(e.get("school")), align=LEFT)
        set_cell(f"C{r}", val(e.get("major")), align=LEFT)
        set_cell(f"D{r}", val(e.get("degree_level")), align=CENTER)
        set_cell(f"E{r}", val(e.get("degree")), align=LEFT)
        r += 1

    positions = data.get("positions") or [PLACEHOLDER]
    pos_start = edu_last_row + 1
    pos_end = pos_start + len(positions) - 1
    merge(f"A{pos_start}:A{pos_end}")
    set_cell(f"A{pos_start}", "現任職位", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    for i, p in enumerate(positions):
        rr = pos_start + i
        merge(f"B{rr}:F{rr}")
        set_cell(f"B{rr}", p, align=LEFT)

    career_rows = data.get("career") or []
    career_header = pos_end + 1
    rows_to_write = career_rows if career_rows else [{}]
    career_end = career_header + len(rows_to_write)
    merge(f"A{career_header}:A{career_end}")
    set_cell(f"A{career_header}", "主要履歷", font=LABEL_FONT, fill=LABEL_FILL, align=CENTER)
    set_cell(f"B{career_header}", "起止年月", font=SUBHEAD_FONT, fill=SUBHEAD_FILL, align=CENTER)
    merge(f"C{career_header}:E{career_header}")
    set_cell(f"C{career_header}", "工作單位", font=SUBHEAD_FONT, fill=SUBHEAD_FILL, align=CENTER)
    set_cell(f"F{career_header}", "職位", font=SUBHEAD_FONT, fill=SUBHEAD_FILL, align=CENTER)
    r = career_header + 1
    for c in rows_to_write:
        set_cell(f"B{r}", val(c.get("date")), align=CENTER)
        merge(f"C{r}:E{r}")
        set_cell(f"C{r}", val(c.get("org")), align=LEFT)
        set_cell(f"F{r}", val(c.get("role")), align=LEFT)
        r += 1

    merge(f"G5:M{career_end}")
    set_cell("G5", "", align=CENTER, border=BORDER_ALL_THIN)

    note_row = career_end + 1
    merge(f"A{note_row}:F{note_row}")
    note_text = data.get("note") or ""
    set_cell(f"A{note_row}", f"注：{note_text}" if note_text else "", font=NOTE_FONT, align=LEFT, border=Border())

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 38.25
    for rr in range(3, career_end + 1):
        ws.row_dimensions[rr].height = 35.1
    ws.row_dimensions[note_row].height = 60
    ws.sheet_view.showGridLines = False

    photos = data.get("photos") or []
    if photos:
        try:
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            print("警告：找不到 Pillow，略過照片嵌入（pip install Pillow 後重跑即可加入照片）")
            photos = []
        anchors = ["G2", "G6"]
        caption_cells = [f"G4", f"G{note_row}"]
        for i, p in enumerate(photos[:2]):
            if not os.path.exists(p["path"]):
                print(f"警告：找不到照片檔案 {p['path']}，略過")
                continue
            img = XLImage(p["path"])
            img.width, img.height = (95, 121) if i == 0 else (130, 168)
            ws.add_image(img, anchors[i])
            cap = ws[caption_cells[i]]
            cap.value = p.get("caption", "").replace("\n", " ") + "（非官方）"
            cap.font = Font(name="Arial", size=9, italic=True, color="808080")

    return wb


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("profile_json", help="輸入 profile.json 路徑")
    ap.add_argument("-o", "--output", required=True, help="輸出 xlsx 路徑")
    args = ap.parse_args()

    with open(args.profile_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    validate_or_exit(data, target="xlsx")

    wb = build(data)
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"wrote {args.output}")


if __name__ == "__main__":
    main()
