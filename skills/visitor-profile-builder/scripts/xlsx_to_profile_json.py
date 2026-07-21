#!/usr/bin/env python3
"""
Extract structured profile data out of a "個人信息登記表" style xlsx
(the format used in reference/*.xlsx) into a profile.json matching
assets/profile.schema.json.

Usage:
    python3 xlsx_to_profile_json.py <input.xlsx> [-o <output.json>]

If -o is omitted, writes <input-basename>.json next to the input file.

Expected xlsx layout (see references/xlsx-source-format.md for the full
reverse-engineered spec):
    A1        timestamp
    A2:F2     title ("個人信息登記表")
    A3/C3     姓名 / 性別 labels, B3/D3:E3 values
    A4/C4     出生年月 / 生肖
    A5/C5     聯繫方式 / 籍貫
    A6:A8+    教育經歷 (merged label) -> header row + N data rows
              (畢業院校 / 專業 / 學歷 / 學位)
    A9:A11+   現任職位 (merged label) -> N rows, each B:F merged
    A12:..+   主要履歷 (merged label) -> header row + N data rows
              (起止年月 / 工作單位(C:E merged) / 職位)
    any cell in column A starting with "注：" -> note
"""
import argparse
import json
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("需要 openpyxl：pip install openpyxl")

PLACEHOLDER_VALUES = {None, "", "-", "－", "—", "–"}
EMPTY = "-"  # the one sanctioned "no data" marker (see validate_profile.py)

# profile_json_to_xlsx.py packs name_en into B3 as 中文名（English）; undo that
# here so the round-trip is lossless. Only splits when the parenthetical is
# Latin text, so a genuine Chinese aside like 王小明（範例）stays in `name`.
NAME_EN_RE = re.compile(r"^(?P<zh>.+?)（(?P<en>[A-Za-z][A-Za-z0-9 .\-']*)）$")

# Likewise, that script writes A1 as 資料彙整：{timestamp}. Strip the label back
# off, or a json -> xlsx -> json round-trip accretes one copy of it per pass.
TS_PREFIX = "資料彙整："


def is_placeholder(v):
    return v is None or str(v).strip() in PLACEHOLDER_VALUES


def cell_value(v):
    """Normalise a cell to the contract: real text, or the literal "-".

    The seed files use several spellings of "empty" (-, －, —, blank); the
    contract allows exactly one, so collapse them all here rather than
    letting an em-dash from a 2019 spreadsheet fail validation downstream."""
    return EMPTY if is_placeholder(v) else str(v).strip()


def strip_ts_label(raw):
    """'資料彙整：2026/01/01' -> '2026/01/01'. Repeats are stripped too, so an
    already-accreted file from before this fix reads back clean."""
    s = str(raw or "").strip()
    while s.startswith(TS_PREFIX):
        s = s[len(TS_PREFIX):].strip()
    return s


def split_name(raw):
    """'王小明（Wang Xiaoming）' -> ('王小明', 'Wang Xiaoming')."""
    raw = (raw or "").strip()
    m = NAME_EN_RE.match(raw)
    if m:
        return m.group("zh").strip(), m.group("en").strip()
    return raw, None


def find_section_rows(ws, label):
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and ws.cell(row=mr.min_row, column=1).value == label:
            return mr.min_row, mr.max_row
    return None


def extract(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    name, name_en = split_name(ws["B3"].value)

    data = {
        "timestamp": strip_ts_label(ws["A1"].value),
        "name": name,
        "gender": cell_value(ws["D3"].value),
        "birth": cell_value(ws["B4"].value),
        "zodiac": cell_value(ws["D4"].value),
        "contact": cell_value(ws["B5"].value),
        "hometown": cell_value(ws["D5"].value),
        "education": [],
        "positions": [],
        "career": [],
        "note": None,
    }

    edu_range = find_section_rows(ws, "教育經歷")
    if edu_range:
        header_row, last_row = edu_range
        for r in range(header_row + 1, last_row + 1):
            school, major, degree_level, degree = (
                ws.cell(row=r, column=c).value for c in (2, 3, 4, 5)
            )
            if not all(is_placeholder(v) for v in (school, major, degree_level, degree)):
                data["education"].append({
                    "school": cell_value(school),
                    "major": cell_value(major),
                    "degree_level": cell_value(degree_level),
                    "degree": cell_value(degree),
                })

    pos_range = find_section_rows(ws, "現任職位")
    if pos_range:
        first_row, last_row = pos_range
        for r in range(first_row, last_row + 1):
            v = ws.cell(row=r, column=2).value
            if not is_placeholder(v):
                data["positions"].append(v)

    career_range = find_section_rows(ws, "主要履歷")
    if career_range:
        header_row, last_row = career_range
        for r in range(header_row + 1, last_row + 1):
            date, org, role = (
                ws.cell(row=r, column=2).value,
                ws.cell(row=r, column=3).value,
                ws.cell(row=r, column=6).value,
            )
            if not all(is_placeholder(v) for v in (date, org, role)):
                data["career"].append({
                    "date": cell_value(date),
                    "org": cell_value(org),
                    "role": cell_value(role),
                })

    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 1 and isinstance(cell.value, str) and cell.value.startswith("注："):
                data["note"] = cell.value[2:].strip()

    if name_en:
        data["name_en"] = name_en

    data["photos"] = []
    data["sources"] = [{"title": "原始來源檔案", "url": ""}]
    return data


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="來源 xlsx 檔路徑")
    ap.add_argument("-o", "--output", help="輸出 json 路徑（預設：同檔名 .json）")
    args = ap.parse_args()

    data = extract(args.input)

    out_path = args.output
    if not out_path:
        import os

        out_path = os.path.splitext(args.input)[0] + ".json"

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
