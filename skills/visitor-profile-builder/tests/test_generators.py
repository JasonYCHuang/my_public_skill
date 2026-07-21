"""Structure of the rendered HTML and xlsx.

No Chrome here — the PNG path can only be verified on a real Linux host, see
scripts/verify-on-ubuntu.sh.
"""
import openpyxl
import pytest

import profile_json_to_xlsx
from profile_json_to_html import render


# --- HTML head --------------------------------------------------------------

def test_html_declares_doctype_charset_and_viewport(profile):
    """Regression: the generated HTML had no <meta> at all. Without charset a
    file:// open can render as mojibake; without viewport the phone layout
    never fires on a real phone."""
    html = render(profile)
    assert html.lstrip().startswith("<!DOCTYPE html>")
    assert '<meta charset="utf-8">' in html
    assert 'name="viewport"' in html


def test_html_title_carries_the_name(profile):
    profile["name"] = "王小明"
    assert "<title>王小明" in render(profile)


def test_font_stack_keeps_a_linux_installable_cjk_family(profile):
    """Regression: the stack listed only macOS/Windows families plus the
    webfont name "Noto Sans TC", which does not match what Ubuntu's
    fonts-noto-cjk registers — so every glyph became a tofu box in the PNG."""
    assert "Noto Sans CJK TC" in render(profile)


# --- 頁尾 -------------------------------------------------------------------

def test_footer_is_dropped_when_no_source_has_a_url(profile):
    """Regression: entry point A seeds sources with an empty url, and url-less
    entries are filtered out — which left an empty grey bar on every
    xlsx-derived card."""
    profile["sources"] = [{"title": "原始來源檔案", "url": ""}]
    # The element, not the string: "card-footer" also appears in the CSS.
    assert '<div class="card-footer">' not in render(profile)


def test_footer_is_present_when_a_source_has_a_url(profile):
    profile["sources"] = [{"title": "官網", "url": "https://example.com"}]
    html = render(profile)
    assert '<div class="card-footer">' in html
    assert "https://example.com" in html


# --- 照片 -------------------------------------------------------------------

def test_placeholder_when_no_photos(profile):
    profile["photos"] = []
    assert "無官方照片" in render(profile)


def test_photo_is_inlined_as_data_uri(profile, photo_file):
    profile["photos"] = [{"path": str(photo_file), "caption": "來源說明"}]
    html = render(profile)
    assert "data:image/png;base64," in html
    assert "來源說明" in html
    assert "無官方照片" not in html


def test_missing_photo_warns_instead_of_raising(profile, capsys):
    """Regression: the html generator raised FileNotFoundError while the xlsx
    generator warned and carried on, so the same profile built one file and
    crashed on the other."""
    profile["photos"] = [{"path": "/nonexistent/nope.jpg", "caption": "x"}]
    html = render(profile)
    assert "找不到照片檔案" in capsys.readouterr().out
    assert "無官方照片" in html


def test_xlsx_missing_photo_also_warns(profile, capsys):
    profile["photos"] = [{"path": "/nonexistent/nope.jpg", "caption": "x"}]
    profile_json_to_xlsx.build(profile)
    assert "找不到照片檔案" in capsys.readouterr().out


# --- 欄位渲染 ---------------------------------------------------------------

def test_all_five_badge_fields_render_even_when_empty(profile):
    for key in ("gender", "birth", "zodiac", "hometown", "contact"):
        profile[key] = "-"
    html = render(profile)
    for label in ("性別", "出生年月", "生肖", "籍貫", "聯繫方式"):
        assert label in html


def test_empty_tables_still_render_a_row(profile):
    profile["education"] = []
    profile["career"] = []
    html = render(profile)
    assert "教育經歷" in html and "主要履歷" in html


def test_html_is_escaped(profile):
    profile["name"] = "<script>alert(1)</script>"
    html = render(profile)
    assert "<script>alert(1)</script>" not in html
    assert "&lt;script&gt;" in html


def test_newline_in_caption_becomes_br(profile, photo_file):
    profile["photos"] = [{"path": str(photo_file), "caption": "第一行\n第二行"}]
    assert "第一行<br>第二行" in render(profile)


# --- xlsx 版面 --------------------------------------------------------------

def test_xlsx_has_the_expected_labels(profile, tmp_path):
    path = tmp_path / "o.xlsx"
    profile_json_to_xlsx.build(profile).save(path)
    ws = openpyxl.load_workbook(path).active
    text = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    for label in ("個人信息登記表", "教育經歷", "現任職位", "主要履歷"):
        assert label in text


def test_xlsx_note_row_is_prefixed(profile, tmp_path):
    profile["note"] = "測試備註"
    path = tmp_path / "o.xlsx"
    profile_json_to_xlsx.build(profile).save(path)
    ws = openpyxl.load_workbook(path).active
    assert any(
        isinstance(c.value, str) and c.value.startswith("注：測試備註")
        for row in ws.iter_rows() for c in row
    )


# --- 兩支產生器都必須先驗證 --------------------------------------------------

@pytest.mark.parametrize("module_name", ["profile_json_to_html", "profile_json_to_xlsx"])
def test_generators_validate_before_writing(module_name, tmp_path, profile, monkeypatch):
    import importlib
    import json
    import sys

    module = importlib.import_module(module_name)
    profile["positions"] = ["a", "b", "c", "d"]      # 超過上限 3
    src = tmp_path / "bad.json"
    src.write_text(json.dumps(profile, ensure_ascii=False), encoding="utf-8")
    out = tmp_path / "out.bin"

    monkeypatch.setattr(sys, "argv", [module_name, str(src), "-o", str(out)])
    with pytest.raises(SystemExit):
        module.main()
    assert not out.exists(), "驗證失敗時不應留下任何檔案"
